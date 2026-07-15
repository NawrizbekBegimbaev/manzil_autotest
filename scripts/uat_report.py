#!/usr/bin/env python3
"""Director-facing UAT report.

Merges the full UAT checklist (docs/testcases/uat/*.json — all 71 cases) with the
latest automated run (allure-results/*-result.json) by case ID, and produces:
  - reports/uat-report-<date>.xlsx  (Сводка по ролям + лист на роль со статусами)
  - a short Russian summary for the director
  - optional Telegram delivery (sendDocument + caption) when TG_* is configured.

Each automated test is titled "<ID> ..." (e.g. "SA-01 ..."), so the run result is
matched to the checklist by the leading ID token. Cases without an automated test
are shown as «Не автоматизирован».

  Build only:  python scripts/uat_report.py --build-only
  Build+send:  python scripts/uat_report.py
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.settings import get_settings  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UAT_DIR = os.path.join(ROOT, "docs", "testcases", "uat")

# checklist file → (role sheet, display role)
ROLES = [
    ("01_super_admin.json", "Super admin", "Супер-администратор"),
    ("02_admin.json", "Admin", "Администратор грузоотправителя"),
    ("03_warehouse.json", "Сотрудник Склада", "Сотрудник склада (мобильное)"),
    ("04_manager.json", "Менеджер", "Менеджер"),
    ("05_tk.json", "TK", "Транспортная компания"),
    ("06_api.json", "API", "API (бэкенд: auth, RBAC)"),
]

RESULT_RU = {"passed": "Пройден", "failed": "Провален", "broken": "Провален",
             "blocked": "Заблокирован (баг)", "skipped": "Пропущен",
             "n/a": "Не автоматизирован"}
RESULT_FILL = {"passed": "C6EFCE", "failed": "FFC7CE", "broken": "FFC7CE",
               "blocked": "F4B183", "skipped": "FFEB9C", "n/a": "E7E6E6"}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _load_run_results(results_dir: str) -> dict:
    """ID (e.g. 'SA-01') -> {status, message}."""
    out: dict[str, dict] = {}
    for path in glob.glob(os.path.join(results_dir, "*-result.json")):
        with open(path, encoding="utf-8") as fh:
            d = json.load(fh)
        name = (d.get("name") or "").strip()
        if not name:
            continue
        cid = name.split(" ", 1)[0].strip().rstrip(".")
        status = d.get("status", "unknown")
        raw_msg = ((d.get("statusDetails") or {}).get("message") or "").strip()
        msg = ""
        if status in ("failed", "broken"):
            msg = raw_msg.splitlines()[0][:300] if raw_msg else ""
        # xfail with a documented bug reason → «Заблокирован», not just «Пропущен».
        if status == "skipped" and ("BUG-" in raw_msg or "Заблокировано" in raw_msg):
            status = "blocked"
            msg = raw_msg.replace("XFAIL", "").replace("[NOTRUN]", "").strip()
            msg = msg.splitlines()[0][:300]
        # keep the worst status if duplicated
        prev = out.get(cid)
        if prev is None or (prev["status"] == "passed" and status != "passed"):
            out[cid] = {"status": status, "message": msg}
    return out


def _hdr(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CTR
    cell.border = BORDER


def build(results_dir: str, out_path: str, env_label: str) -> dict:
    run = _load_run_results(results_dir)
    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"

    per_role = []  # (role, total, automated, passed, failed)
    roles_data = []
    for fname, sheet, role in ROLES:
        cases = json.load(open(os.path.join(UAT_DIR, fname), encoding="utf-8"))
        rows = []
        total = len(cases)
        automated = passed = failed = blocked = 0
        for c in cases:
            cid = c["id"]
            r = run.get(cid)
            status = r["status"] if r else "n/a"
            if r:
                automated += 1
                if status == "passed":
                    passed += 1
                elif status in ("failed", "broken"):
                    failed += 1
                elif status == "blocked":
                    blocked += 1
            rows.append({**c, "_status": status, "_msg": r["message"] if r else ""})
        per_role.append((role, total, automated, passed, failed, blocked))
        roles_data.append((sheet, role, rows))

    # ── Сводка sheet ──
    today = datetime.now().strftime("%d.%m.%Y")
    summary.merge_cells("A1:G1")
    t = summary["A1"]
    t.value = f"Manzil — отчёт автотестов UAT   ·   окружение: {env_label}   ·   дата: {today}"
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="left", vertical="center")
    summary.row_dimensions[1].height = 26

    cols = ["Роль", "Всего сценариев", "Автоматизировано", "Пройдено", "Провалено",
            "Заблокировано (баг)", "% прохождения", "% автоматизации"]
    summary.append(cols)
    for c in range(1, len(cols) + 1):
        _hdr(summary.cell(row=2, column=c))
    widths = [34, 15, 17, 11, 11, 18, 15, 16]
    for i, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(i)].width = w

    T = A = P = F = B = 0
    for role, total, automated, passed, failed, blocked in per_role:
        T += total; A += automated; P += passed; F += failed; B += blocked
        testable = automated - blocked  # blocked = заблокировано багом, не считаем в знаменателе
        pct_pass = f"{round(100 * passed / testable)}%" if testable else "—"
        pct_auto = f"{round(100 * automated / total)}%" if total else "—"
        summary.append([role, total, automated, passed, failed, blocked, pct_pass, pct_auto])
        row = summary.max_row
        for c in range(1, len(cols) + 1):
            cell = summary.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CTR if c > 1 else WRAP
        if failed:
            summary.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FFC7CE")
        if blocked:
            summary.cell(row=row, column=6).fill = PatternFill("solid", fgColor="F4B183")

    summary.append(["ИТОГО", T, A, P, F, B,
                    f"{round(100 * P / (A - B))}%" if (A - B) else "—",
                    f"{round(100 * A / T)}%" if T else "—"])
    trow = summary.max_row
    for c in range(1, len(cols) + 1):
        cell = summary.cell(row=trow, column=c)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = CTR if c > 1 else WRAP
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    # ── per-role sheets ──
    failures = []
    for sheet, role, rows in roles_data:
        ws = wb.create_sheet(title=sheet[:31])
        cols2 = ["ID", "Экран", "Сценарии", "Приоритет", "Результат", "Примечание"]
        ws.append(cols2)
        for c in range(1, len(cols2) + 1):
            _hdr(ws.cell(row=1, column=c))
        for i, w in enumerate([10, 26, 44, 12, 20, 50], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        for r in rows:
            st = r["_status"]
            ws.append([r["id"], r.get("screen", ""), r.get("scenario", ""),
                       r.get("priority", ""), RESULT_RU.get(st, st), r.get("_msg", "")])
            row = ws.max_row
            for c in range(1, len(cols2) + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = BORDER
                cell.alignment = CTR if c in (1, 4, 5) else WRAP
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=RESULT_FILL.get(st, "FFFFFF"))
            if st in ("failed", "broken"):
                failures.append((r["id"], r.get("scenario", ""), r.get("_msg", "")))

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)

    # ── director text summary ──
    lines = [f"📋 Manzil — автотесты UAT [{env_label}] · {today}",
             f"Пройдено: {P} из {A - B} проверяемых "
             f"(ещё {B} заблокировано багом; всего сценариев в чеклисте: {T}).",
             ""]
    lines.append("По ролям (пройдено / проверяемо / всего, +заблокировано):")
    for role, total, automated, passed, failed, blocked in per_role:
        testable = automated - blocked
        mark = "✅" if failed == 0 and testable else ("⚠️" if failed else "—")
        extra = f" (+{blocked} забл.)" if blocked else ""
        lines.append(f"  {mark} {role}: {passed}/{testable}/{total}{extra}")
    if failures:
        lines.append("")
        lines.append(f"❌ Провалено ({len(failures)}):")
        for cid, scen, msg in failures[:15]:
            lines.append(f"  • {cid} {scen}" + (f" — {msg}" if msg else ""))
    else:
        lines.append("")
        lines.append("Провалов нет среди автоматизированных сценариев.")
    caption = "\n".join(lines)
    return {"out_path": out_path, "caption": caption, "passed": P,
            "automated": A, "total": T, "failures": failures}


def _send_telegram(out_path, caption, token, chat_id):
    import requests
    with open(out_path, "rb") as fh:
        resp = requests.post(
            f"https://api.telegram.org/bot{token}/sendDocument",
            data={"chat_id": chat_id, "caption": caption[:1024]},
            files={"document": (os.path.basename(out_path), fh)}, timeout=60)
    if resp.status_code != 200:
        print(f"Telegram send failed: {resp.status_code} {resp.text}", file=sys.stderr)
        return False
    print("Telegram: отчёт отправлен директору.")
    return True


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--results-dir", default="allure-results")
    ap.add_argument("--out", default=None)
    ap.add_argument("--build-only", action="store_true", help="Не отправлять в Telegram")
    args = ap.parse_args()

    cfg = get_settings()
    out = args.out or os.path.join(ROOT, "reports",
                                   f"uat-report-{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    res = build(args.results_dir, out, cfg.tg_report_env)
    print(f"XLSX: {res['out_path']}")
    print(res["caption"])

    if args.build_only:
        return 0
    if not (cfg.tg_bot_token and cfg.tg_chat_id):
        print("\nTelegram не настроен (.env TG_BOT_TOKEN/TG_CHAT_ID) — отправка пропущена.")
        return 0
    _send_telegram(res["out_path"], res["caption"], cfg.tg_bot_token, cfg.tg_chat_id)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
