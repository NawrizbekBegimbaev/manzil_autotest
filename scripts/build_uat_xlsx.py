#!/usr/bin/env python3
"""Build the business-readable UAT checklist (by role) from docs/testcases/uat/*.json.

Sheets: Сводка → Super admin → Admin → Сотрудник Склада → Менеджер → TK.
Two-row header: single columns ID | Экран | Сценарии | Шаги | Приоритет |
Ожидаемый результат span both rows; "Фактический результат" merges over two
sub-columns "QA" and "Пользователь". Plain language, positive scenarios only.
"""

from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "docs", "testcases", "uat")
OUT = os.path.join(ROOT, "docs", "testcases", "manzil-uat-checklist.xlsx")

AREAS = [
    ("01_super_admin.json", "Super admin", "Супер-администратор",
     "Вход, создание грузоотправителей и транспортных компаний, водители, справочники (города, типы транспорта)"),
    ("02_admin.json", "Admin", "Администратор грузоотправителя",
     "Вход, дашборд, заказы и их отклики, выбор перевозчика, сотрудники, склады, отчёты, отмена/завершение"),
    ("03_warehouse.json", "Сотрудник Склада", "Сотрудник склада (мобильное приложение)",
     "Вход, список заявок, создание заявки и публикация, связь с водителем, отправка груза"),
    ("04_manager.json", "Менеджер", "Менеджер (оператор склада / диспетчер)",
     "Вход, оператор склада, диспетчер, связь с водителем, отмена и повторная публикация"),
    ("05_tk.json", "TK", "Транспортная компания (перевозчик)",
     "Вход, лента заявок, предложить/изменить цену, мои предложения, водители, назначение и старт заявки"),
]

# single-row columns (merged vertically over header rows 1-2)
SINGLE_COLS = ["ID", "Экран", "Сценарии", "Шаги", "Приоритет", "Ожидаемый результат"]
WIDTHS = [10, 26, 40, 58, 12, 50, 16, 18]  # last two = QA / Пользователь
FIELDS = ["id", "screen", "scenario", "steps", "priority", "expected", "_qa", "_user"]
NCOLS = 8

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PRIO_FILL = {"Высокий": "FFC7CE", "Средний": "FFEB9C", "Низкий": "E2EFDA"}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _hdr(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def _build_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name[:31])
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Single columns A..F merged across rows 1-2
    for c, name in enumerate(SINGLE_COLS, start=1):
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
        ws.cell(row=1, column=c, value=name)
    # "Фактический результат" merged over G1:H1, sub-headers in row 2
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    ws.cell(row=1, column=7, value="Фактический результат")
    ws.cell(row=2, column=7, value="QA")
    ws.cell(row=2, column=8, value="Пользователь")
    # style every header cell in rows 1-2
    for r in (1, 2):
        for c in range(1, NCOLS + 1):
            _hdr(ws.cell(row=r, column=c))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for r in rows:
        ws.append([r.get(f, "") for f in FIELDS])
        row = ws.max_row
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 5, 7, 8) else WRAP_TOP
        prio = r.get("priority", "")
        if prio in PRIO_FILL:
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=PRIO_FILL[prio])
    return ws


def _build_index(wb, counts, total):
    idx = wb.active
    idx.title = "Сводка"
    cols = ["Роль", "Лист", "Что проверяем (простыми словами)", "Кол-во сценариев"]
    idx.append(cols)
    for i, w in enumerate([34, 22, 78, 18], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, len(cols) + 1):
        _hdr(idx.cell(row=1, column=c))
    idx.row_dimensions[1].height = 26
    idx.freeze_panes = "A2"

    for role, sheet, about, n in counts:
        idx.append([role, sheet, about, n])
        row = idx.max_row
        for c in range(1, len(cols) + 1):
            cell = idx.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c == 4 else WRAP_TOP

    idx.append(["ИТОГО", "", "", total])
    trow = idx.max_row
    for c in range(1, len(cols) + 1):
        cell = idx.cell(row=trow, column=c)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = CENTER if c == 4 else WRAP_TOP
        cell.fill = PatternFill("solid", fgColor="E2EFDA")

    note_row = idx.max_row + 2
    idx.cell(row=note_row, column=1,
             value=("Как пользоваться: откройте лист своей роли и выполняйте сценарии по шагам сверху вниз. "
                    "Если результат совпал с «Ожидаемым» — пишите «ОК» (Passed) в колонке «Фактический результат». "
                    "Если нет — опишите, что произошло. Подколонки: QA заполняет тестировщик, Пользователь — заказчик."))
    idx.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    nc = idx.cell(row=note_row, column=1)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    nc.font = Font(italic=True)
    idx.row_dimensions[note_row].height = 60


def main() -> int:
    wb = Workbook()
    counts = []
    total = 0
    # build area sheets first (active sheet becomes index after we build it)
    for fname, sheet, role, about in AREAS:
        rows = json.load(open(os.path.join(DATA, fname), encoding="utf-8"))
        total += len(rows)
        counts.append((role, sheet, about, len(rows)))
        _build_sheet(wb, sheet, rows)
    _build_index(wb, counts, total)  # renames the default sheet → "Сводка" (stays first)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total scenarios: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
