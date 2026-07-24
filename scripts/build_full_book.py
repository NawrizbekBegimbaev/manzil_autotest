#!/usr/bin/env python3
"""Assemble ONE combined workbook from the current test-case source (all 1497):
docs/testcases/{api,web,mobile}/*.json → docs/testcases/manzil-testcases-full.xlsx.

Same layout/styling as the per-area books: a "Сводка" index sheet + one sheet per
section with columns ID | Раздел | Экран/Подраздел | Сценарий | Предусловия | Шаги |
Ожидаемый результат | Приоритет | Статус | Комментарий.
"""

from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TC = os.path.join(ROOT, "docs", "testcases")
OUT = os.path.join(TC, "manzil-testcases-full.xlsx")

# (subdir, file, sheet ≤31, display раздел, маршруты/охват)
SECTIONS = [
    ("api", "01_auth_me_devices.json", "API — Аутентификация", "API — Auth/Me/Устройства",
     "/auth/login · /auth/refresh · /auth/logout · /me · /me/devices · clientType"),
    ("api", "02_superadmin.json", "API — Супер-админ", "API — SUPER_ADMIN",
     "/super-admin/shipper-companies · transport-companies · drivers · cities · vehicle-types"),
    ("api", "03_shipper_orders_staff.json", "API — Грузоотправитель", "API — Shipper (заказы/персонал)",
     "/shipper/orders · offers · staff · warehouses · reports · dashboard"),
    ("api", "04_tendering_transport.json", "API — Тендер+Перевозчик", "API — Тендер / TRANSPORT",
     "/transport/feed · offers · drivers · assignment · start · winner-select"),
    ("api", "05_warehouse_dispatch.json", "API — Склад+Диспетчер", "API — Warehouse/Dispatch",
     "/warehouse/orders · locations · communication · goods-sent · departures · dispatch-log"),
    ("api", "06_integrations_sms_dicts.json", "API — Интеграции+SMS", "API — Интеграции/SMS/Справочники",
     "/integrations/1c · /shipper/sms · sms-logs · countries · cities · cn/divisions"),
    ("api", "07_rbac_capabilities.json", "API — RBAC", "API — RBAC и capability",
     "Роли · права (capabilities) · tenancy · wrong-app · 401/403/404"),
    ("api", "08_registration_reset_public.json", "API — Регистрация", "API — Регистрация/Сброс/Публичное",
     "/auth/register/* (OTP) · /auth/reset/* · transporters/{id}/verification · /public/cargo · публичные справочники"),
    ("web", "01_auth_shell_settings.json", "Web — Auth+Оболочка", "Web — Auth/Оболочка/Настройки",
     "Логин · навигация · язык (China-first) · профиль · настройки"),
    ("web", "02_superadmin.json", "Web — Супер-админ", "Web — SUPER_ADMIN",
     "Компании · водители · справочники (города/дивизионы/типы ТС)"),
    ("web", "03_shipper.json", "Web — Грузоотправитель", "Web — Shipper (ADMIN/MANAGER)",
     "Список заказов · оффер/выбор · персонал · склады · отчёты · диспетчер"),
    ("web", "04_transport.json", "Web — Перевозчик", "Web — TRANSPORT",
     "Лента заявок · предложение цены · водители · назначение/старт"),
    ("web", "05_rbac_archive.json", "Web — RBAC+Архив", "Web — RBAC / Архив",
     "Ролевые редиректы · доступность действий · архивные состояния"),
    ("mobile", "01_warehouse_app.json", "Мобайл — Склад", "Мобайл — Warehouse-app",
     "Логин · создание заявки · адрес (cityId) · публикация · lifecycle · связь · goods-sent"),
    ("mobile", "02_carrier_app.json", "Мобайл — Перевозчик", "Мобайл — Carrier-app",
     "Логин · лента · предложение · водители · назначение · старт заявки"),
]

COLUMNS = ["ID", "Раздел", "Экран / Подраздел", "Сценарий", "Предусловия",
           "Шаги", "Ожидаемый результат", "Приоритет", "Статус", "Комментарий"]
WIDTHS = [14, 26, 26, 42, 34, 52, 56, 12, 10, 16]
FIELDS = ["id", "razdel", "screen", "scenario", "precond", "steps", "expected",
          "priority", "_status", "comment"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PRIO_FILL = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "E2EFDA"}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _build_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append(COLUMNS)
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws, len(COLUMNS))
    for r in rows:
        r = {**r, "_status": ""}
        ws.append([r.get(f, "") for f in FIELDS])
        row = ws.max_row
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 8, 9) else WRAP_TOP
        prio = r.get("priority", "")
        if prio in PRIO_FILL:
            ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=PRIO_FILL[prio])
    return ws


def main() -> int:
    wb = Workbook()
    idx = wb.active
    idx.title = "Сводка"
    idx_cols = ["Раздел", "Лист", "Маршруты / охват", "Слой", "Кейсов"]
    idx.append(idx_cols)
    for i, w in enumerate([34, 26, 80, 12, 10], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w
    _style_header(idx, len(idx_cols))

    total = 0
    by_prio = {"High": 0, "Medium": 0, "Low": 0}
    by_layer = {"api": 0, "web": 0, "mobile": 0}
    seen_ids = set()
    dups = []
    for subdir, fname, sheet, razdel, routes in SECTIONS:
        rows = json.load(open(os.path.join(TC, subdir, fname), encoding="utf-8"))
        for r in rows:
            cid = r.get("id", "")
            if cid in seen_ids:
                dups.append(cid)
            seen_ids.add(cid)
            by_prio[r.get("priority", "")] = by_prio.get(r.get("priority", ""), 0) + 1
        total += len(rows)
        by_layer[subdir] += len(rows)
        _build_sheet(wb, sheet, rows)
        idx.append([razdel, sheet[:31], routes, subdir.upper(), len(rows)])
        row = idx.max_row
        for c in range(1, len(idx_cols) + 1):
            cell = idx.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c in (4, 5) else WRAP_TOP

    idx.append(["ИТОГО", "", f"API {by_layer['api']} · Web {by_layer['web']} · Mobile {by_layer['mobile']}",
                "", total])
    trow = idx.max_row
    for c in range(1, len(idx_cols) + 1):
        cell = idx.cell(row=trow, column=c)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = CENTER if c in (4, 5) else WRAP_TOP
    idx.cell(row=trow, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    idx.cell(row=trow, column=5).fill = PatternFill("solid", fgColor="DDEBF7")

    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Total: {total}  (API {by_layer['api']} / Web {by_layer['web']} / Mobile {by_layer['mobile']})")
    print(f"Priority: High {by_prio['High']} / Medium {by_prio['Medium']} / Low {by_prio['Low']}")
    print(f"Unique IDs: {len(seen_ids)} | duplicates: {dups or 'none'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
