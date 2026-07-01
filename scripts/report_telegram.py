#!/usr/bin/env python3
"""Build an XLSX sanity report from allure-results and (optionally) send it to Telegram.

The Allure result JSONs already carry human-readable titles ("N. Case"), set via
`@allure.title`. We read them, build an XLSX (№ | Кейс | Статус | Примечание) with
a colored status cell and a summary banner, then — unless --build-only — send the
file to Telegram via the Bot API.

Telegram config comes from .env (TG_BOT_TOKEN, TG_CHAT_ID, TG_REPORT_ENV). If the
token/chat are empty, sending is skipped silently but the XLSX is still written.

  Build only (used by run_sanity.sh):  python scripts/report_telegram.py --build-only
  Build + send (manual, once a day):   python scripts/report_telegram.py
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.settings import get_settings  # noqa: E402

# Status → fill color (ARGB) for the status cell.
STATUS_FILL = {
    "passed": "C6EFCE",
    "failed": "FFC7CE",
    "broken": "FFEB9C",
    "skipped": "D9D9D9",
    "unknown": "FFFFFF",
}
STATUS_RU = {
    "passed": "Пройден",
    "failed": "Провален",
    "broken": "Сломан",
    "skipped": "Пропущен",
    "unknown": "—",
}
_SORT_KEY = {"failed": 0, "broken": 1, "passed": 2, "skipped": 3, "unknown": 4}


def _read_results(results_dir: str) -> list[dict]:
    rows: list[dict] = []
    for path in glob.glob(os.path.join(results_dir, "*-result.json")):
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        title = data.get("name") or data.get("fullName") or "(без названия)"
        status = data.get("status", "unknown")
        note = ""
        if status in ("failed", "broken"):
            note = (data.get("statusDetails") or {}).get("message", "") or ""
            note = note.strip().splitlines()[0][:300] if note else ""
        rows.append({"title": title.strip(), "status": status, "note": note})

    def order(r: dict):
        title = r["title"]
        num = title.split(".", 1)[0]
        return (int(num) if num.isdigit() else 9999, title)

    rows.sort(key=order)
    return rows


def _build_xlsx(rows: list[dict], out_path: str, env_label: str) -> dict:
    counts = {k: 0 for k in STATUS_FILL}
    for r in rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    total = len(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sanity"

    banner = (
        f"Manzil Sanity [{env_label}] — всего: {total}, "
        f"пройдено: {counts.get('passed', 0)}, провалено: {counts.get('failed', 0)}, "
        f"сломано: {counts.get('broken', 0)}, пропущено: {counts.get('skipped', 0)}"
    )
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = banner
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["№", "Кейс", "Статус", "Примечание"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")

    for i, r in enumerate(rows, start=1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r["title"])
        st = ws.cell(row=row, column=3, value=STATUS_RU.get(r["status"], r["status"]))
        st.fill = PatternFill("solid", fgColor=STATUS_FILL.get(r["status"], "FFFFFF"))
        st.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=r["note"])

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 70

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return {"total": total, "counts": counts, "banner": banner}


def _send_telegram(out_path: str, caption: str, token: str, chat_id: str) -> bool:
    import requests

    with open(out_path, "rb") as fh:
        resp = requests.post(
            f"https://api.telegram.org/bot{token}/sendDocument",
            data={"chat_id": chat_id, "caption": caption},
            files={"document": (os.path.basename(out_path), fh)},
            timeout=60,
        )
    if resp.status_code != 200:
        print(f"Telegram send failed: {resp.status_code} {resp.text}", file=sys.stderr)
        return False
    print("Telegram: report sent.")
    return True


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--results-dir", default="allure-results")
    ap.add_argument("--out", default="reports/sanity-report.xlsx")
    ap.add_argument("--build-only", action="store_true", help="Build XLSX, do not send")
    args = ap.parse_args()

    cfg = get_settings()
    rows = _read_results(args.results_dir)
    if not rows:
        print(f"No *-result.json in {args.results_dir}", file=sys.stderr)
        return 1

    summary = _build_xlsx(rows, args.out, cfg.tg_report_env)
    print(f"XLSX written: {args.out}")
    print(summary["banner"])

    if args.build_only:
        return 0
    if not (cfg.tg_bot_token and cfg.tg_chat_id):
        print("Telegram not configured (.env) — send skipped.")
        return 0
    _send_telegram(args.out, summary["banner"], cfg.tg_bot_token, cfg.tg_chat_id)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
