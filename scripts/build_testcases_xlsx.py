#!/usr/bin/env python3
"""Assemble the Manzil test-case workbook (XLSX) from docs/testcases/data/*.json.

Mirrors the MyXodim/BusinessHub layout: a "Сводка" index sheet + one sheet per
area with columns: ID | Раздел | Экран / Подраздел | Сценарий | Предусловия |
Шаги | Ожидаемый результат | Приоритет | Статус | Комментарий.
"""

from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "docs", "testcases", "data")
OUT = os.path.join(ROOT, "docs", "testcases", "manzil-testcases.xlsx")

# file → (sheet name ≤31 chars, display раздел, маршруты/охват)
AREAS = [
    ("01_auth.json", "Аутентификация", "Аутентификация и сессии",
     "/auth/login · /auth/refresh · /auth/logout · /me · clientType (WEB/WAREHOUSE_APP/TRANSPORT_COMPANY_APP)"),
    ("02_rbac.json", "RBAC", "Доступы и роли",
     "Матрица ролей · 401/403/404 · blocked-company · wrong-app · web silent-redirect"),
    ("03_superadmin_companies.json", "SA — Компании", "SUPER_ADMIN — Компании",
     "/super-admin/shipper-companies · /super-admin/transport-companies"),
    ("04_superadmin_drivers_dicts.json", "SA — Водители+Справочники", "SUPER_ADMIN — Водители/Справочники",
     "/super-admin/drivers · /super-admin/cities · /super-admin/vehicle-types"),
    ("05_shipper.json", "Грузоотправитель", "Грузоотправитель (ADMIN/MANAGER)",
     "/shipper/orders · /shipper/staff · /shipper/warehouses · /shipper/reports · /dashboard"),
    ("06_carrier_tender.json", "Перевозчик+Тендер", "Перевозчик (TRANSPORT_ADMIN) и Тендер",
     "/transport/feed · offers · drivers · assignment · start · winner-select"),
    ("07_mobile.json", "Мобайл", "Мобильные приложения",
     "Warehouse-app (создание заявки, lifecycle) · Carrier-app (feed, offer, drivers, start)"),
    ("08_integrations_e2e.json", "Интеграции+E2E", "Интеграции, Уведомления, Файлы, E2E",
     "/integrations/1c · /me/notifications · /me/devices · /files · сквозной тендер E2E"),
]

COLUMNS = ["ID", "Раздел", "Экран / Подраздел", "Сценарий", "Предусловия",
           "Шаги", "Ожидаемый результат", "Приоритет", "Статус", "Комментарий"]
WIDTHS = [12, 24, 26, 42, 34, 52, 56, 12, 10, 14]
FIELDS = ["id", "razdel", "screen", "scenario", "precond", "steps", "expected",
          "priority", "_status", "comment"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PRIO_FILL = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "E2EFDA"}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _build_area_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append(COLUMNS)
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws, len(COLUMNS))
    for r in rows:
        r = {**r, "_status": ""}
        ws.append([r.get(f, "") for f in FIELDS])
        row = ws.max_row
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 8, 9) else WRAP_TOP
        prio = r.get("priority", "")
        if prio in PRIO_FILL:
            ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=PRIO_FILL[prio])
    return ws


def main() -> int:
    wb = Workbook()
    # Index sheet
    idx = wb.active
    idx.title = "Сводка"
    idx_cols = ["Раздел", "Лист", "Маршруты / охват", "Покрытие", "Кейсов"]
    idx.append(idx_cols)
    for i, w in enumerate([34, 26, 80, 14, 10], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w
    _style_header(idx, len(idx_cols))

    total = 0
    by_prio = {"High": 0, "Medium": 0, "Low": 0}
    for fname, sheet, razdel, routes in AREAS:
        rows = json.load(open(os.path.join(DATA, fname), encoding="utf-8"))
        total += len(rows)
        for r in rows:
            by_prio[r.get("priority", "")] = by_prio.get(r.get("priority", ""), 0) + 1
        _build_area_sheet(wb, sheet, rows)
        idx.append([razdel, sheet[:31], routes, "Готово", len(rows)])
        row = idx.max_row
        for c in range(1, len(idx_cols) + 1):
            cell = idx.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c in (4, 5) else WRAP_TOP

    # total row
    idx.append(["ИТОГО", "", "", "", total])
    trow = idx.max_row
    for c in range(1, len(idx_cols) + 1):
        cell = idx.cell(row=trow, column=c)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = CENTER if c in (4, 5) else WRAP_TOP
    idx.cell(row=trow, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    idx.cell(row=trow, column=5).fill = PatternFill("solid", fgColor="DDEBF7")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Total cases: {total} | High {by_prio['High']} / Medium {by_prio['Medium']} / Low {by_prio['Low']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
