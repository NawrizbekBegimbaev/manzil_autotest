#!/usr/bin/env python3
"""Assemble the three per-layer workbooks from docs/testcases/{api,web,mobile}/*.json,
reusing the layout/styling of build_full_book.py (same sheets + «Сводка» index):

  api    → docs/testcases/manzil-api-testcases.xlsx
  web    → docs/testcases/manzil-web-testcases.xlsx
  mobile → docs/testcases/manzil-mobile-testcases.xlsx
"""

from __future__ import annotations

import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import build_full_book as B

OUTS = {
    "api": "manzil-api-testcases.xlsx",
    "web": "manzil-web-testcases.xlsx",
    "mobile": "manzil-mobile-testcases.xlsx",
}


def build_layer(layer: str) -> int:
    wb = Workbook()
    idx = wb.active
    idx.title = "Сводка"
    idx_cols = ["Раздел", "Лист", "Маршруты / охват", "Слой", "Кейсов"]
    idx.append(idx_cols)
    for i, w in enumerate([34, 26, 80, 12, 10], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w
    B._style_header(idx, len(idx_cols))

    total = 0
    for subdir, fname, sheet, razdel, routes in B.SECTIONS:
        if subdir != layer:
            continue
        rows = json.load(open(os.path.join(B.TC, subdir, fname), encoding="utf-8"))
        total += len(rows)
        B._build_sheet(wb, sheet, rows)
        idx.append([razdel, sheet[:31], routes, subdir.upper(), len(rows)])
        row = idx.max_row
        for c in range(1, len(idx_cols) + 1):
            cell = idx.cell(row=row, column=c)
            cell.border = B.BORDER
            cell.alignment = B.CENTER if c in (4, 5) else B.WRAP_TOP

    idx.append(["ИТОГО", "", "", layer.upper(), total])
    trow = idx.max_row
    for c in range(1, len(idx_cols) + 1):
        cell = idx.cell(row=trow, column=c)
        cell.font = Font(bold=True)
        cell.border = B.BORDER
        cell.alignment = B.CENTER if c in (4, 5) else B.WRAP_TOP
    idx.cell(row=trow, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    idx.cell(row=trow, column=5).fill = PatternFill("solid", fgColor="DDEBF7")

    out = os.path.join(B.TC, OUTS[layer])
    wb.save(out)
    print(f"Saved {out} ({total} cases)")
    return total


def main() -> int:
    grand = sum(build_layer(layer) for layer in ("api", "web", "mobile"))
    print(f"Grand total across layer books: {grand}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
